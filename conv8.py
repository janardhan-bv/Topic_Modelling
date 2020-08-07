import pandas as pd
import openpyxl
import nltk
from openpyxl import load_workbook
from tika import parser
nltk.download('wordnet')
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
from nltk.corpus import brown

nltk.download('brown')

raw = parser.from_file('C://Users//bvjan//Documents//data1.pdf')
my = raw['content']
print(my)

words = nltk.word_tokenize(my)
t1 = nltk.pos_tag(words)
print(t1)

train = {}
train['tag'] = t1
t5 = ([t[0] for t in train['tag'] if ((t[1] == 'NN') or (t[1] == 'NNS') or (t[1] == 'VBG') or (t[1] == 'VBD'))])
print(t5)
for t in t5:
    if not (len(t) > 2):
        t5.remove(t)
print(t5)


# ********************************************************************************
import xlsxwriter

workbook = xlsxwriter.Workbook('Example2.xlsx')
worksheet = workbook.add_worksheet()

# Start from the first cell.
# Rows and columns are zero indexed.
row = 0
column = 0

content = t5
# iterating through content list
for item in content:
    # write operation perform
    worksheet.write(row, column, item)

    # incrementing the value of row by one
    # with each iteratons.
    row += 1

# ****************************************************************************************************
from nltk.corpus import wordnet
synonyms = []

row = 0

for t in t5:
    column = 1
    for syn in wordnet.synsets(t):
        for l in syn.lemmas():
            synonyms.append(l.name())
            worksheet.write(row, column, l.name())
            column += 1
    row += 1

workbook.close()
# ********************************************************************
a = []
book = openpyxl.load_workbook('Example2.xlsx')
sheet = book.sheetnames
currentsheet = book[sheet[0]]
for i in range(0, currentsheet.max_row):
    count = 1
    for row in range(i+2, currentsheet.max_row+1):
        for column in "ABCDEFGHIJLMNOPQRST":
            cell_name = "{}{}".format(column, row)
            if t5[i] == currentsheet[cell_name].value:
               count +=1
               break
    a.append(count)

# *************************************************************************
workbook1 = xlsxwriter.Workbook('Example3.xlsx')
worksheet1 = workbook1.add_worksheet()
name = 'name'
number = 'number'
prob = 'probablity'
worksheet1.write('A1', name)
worksheet1.write('B1', number)
worksheet1.write('C1', prob)
row = 1
column = 0
x = len(t5)

# iterating through content list
for item in t5:
    worksheet1.write(row, column, item)
    worksheet1.write(row,column+1,a[row-1])
    worksheet1.write(row, column+2, a[row-1]/x)
    # incrementing the value of row by one
    # with each iteratons.
    row += 1
workbook1.close()

# *******************************************************************************
import pandas as pd

file_df = pd.read_excel("Example3.xlsx")

# Keep only FIRST record from set of duplicates
file_df_first_record = file_df.drop_duplicates(subset=["name"], keep="first")
file_df_first_record.to_excel("Records.xlsx", index=False)

# ******************************************************************************
# working
import gensim
from gensim import corpora
import numpy as np

# array = np.array(t5)
# print(array)
dataset = [d.split() for d in t5]
dictionary = corpora.Dictionary(dataset)
doc_term_matrix = [dictionary.doc2bow(doc) for doc in dataset]
print(doc_term_matrix)

Lda = gensim.models.ldamodel.LdaModel
ldamodel = Lda(doc_term_matrix, num_topics=3, id2word = dictionary, passes=50)
print(ldamodel.print_topics(num_topics=3, num_words=3))

# ********************************************************************************************

import tomotopy as tp

mdl = tp.CTModel(k=20)
for line in open('C://Users//bvjan//Documents//data1.txt'):
    mdl.add_doc(line.strip().split())

for i in range(0, 100, 10):
    mdl.train(10)
    print('Iteration: {}\tLog-likelihood: {}'.format(i, mdl.ll_per_word))

for k in range(mdl.k):
    print('Top 10 words of topic #{}'.format(k))
    print(mdl.get_topic_words(k, top_n=10))
# **********************************************************************************************
